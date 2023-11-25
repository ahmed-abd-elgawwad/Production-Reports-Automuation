import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings("ignore")


# ------------------------------------ Fetching Functions -------------------------------------------------

class OneFileFetch:

    def __init__(self, file_path):

        self.file_path = file_path

        # determine name of th wells
        self.Wells_dic = [['M-1', 'Muzhil-1'], ['M-2', 'Muzhil-2'], ['M-3D', 'Muzhil-3D'],['M-3B', 'Muzhil-3B'], ['M-5', 'Muzhil-5'], ['M-9', 'Muzhil-9']]
        self.Wells = pd.DataFrame(self.Wells_dic, columns=['Short_Name', 'Well'])

    def handle_hours_format(self, hour):
    
        if isinstance(hour, (float, int)):
            # Handle float values (e.g., 5.5)
            return hour

        elif isinstance(hour, str):
            # Handle time values in the format 'HH:MM:SS' (e.g., '23:48:00')
            try:
                if ":" in hour:
                    time_parts = hour.split(':')

                    # if there is  [ hours : minutes : seconds ]
                    if len(time_parts) == 3:
                        hour = int(time_parts[0]) + (int(time_parts[1]) / 60) + (int(time_parts[2]) / 3600)
                        return hour

                    # if there is  [ hours : minutes]
                    elif len(time_parts) == 2:
                        hour = int(time_parts[0]) + (int(time_parts[1]) / 60) 
                        return hour
                else:
                    return hour

            except ValueError:
                return hour
        else:
            return 24
        

    def get_tables_ranges(self,sheet):
    
        # words taht determines the start and end row for the tables
        search_words = ['Wells',"Total","Name", 'Cumulative', 'Storage Tanks',"Injection Rate, BPD","Freq.","Slop 1"]

        # Create a dictionary to store indices for each search word
        word_indices = {word: [] for word in search_words}

        # Iterate through rows to search for the words in the first column
        for row_number, row in enumerate(sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
            cell_value = row[0]
            for word in search_words[:-3]:
                if cell_value == word:
                    word_indices[word].append(row_number)
                    
        # Iterate through rows to search for the words in the first column
        for row_number, row in enumerate(sheet.iter_rows(min_col=9, max_col=9, values_only=True), start=1):
            cell_value = row[0]
            for word in (search_words[-3],search_words[-2]) :
                if cell_value == word:
                    word_indices[word].append(row_number)
                    
        # Iterate through rows to search for the words in the first column
        for row_number, row in enumerate(sheet.iter_rows(min_col=4, max_col=4, values_only=True), start=1):
            cell_value = row[0]
            if cell_value == search_words[-1]:
                word_indices[search_words[-1]].append(row_number)
                    
        return word_indices

    def fetch_hours_data(self,ws, indices):
            
        start = indices["Wells"][-1]
        end = start +6
        
        # Get the table correspoding the production hours data 
        table_range = ws[f"A{start}" : f"E{end}" ]
        rows_hours= [[value.value for value   in row] for row in table_range]
        wells_short_name = [row[0] for row in rows_hours[1:]]
        status = [row[1] for row in rows_hours[1:]]
        hours = [row[2] for row in rows_hours[1:]]
        reasons = [row[-1] for row in rows_hours[1:]]
        d = pd.DataFrame()
        d[0] = wells_short_name
        d[0] = d[0].astype(str).str.strip()
        d["hours"] = hours
        d["status"] = status
        d["status"] = d["status"].fillna(method="ffill")
        d["reason"] = reasons

        # d = d[d[0].isin(Wells_19["Short_Name"])]
        d[0] = d[0].replace({ short_name : name  for short_name , name in self.Wells_dic})
        
        return d

    def fetch_production_data(self , ws , indcies):

        start = indcies["Wells"][0]
        end = indcies["Total"][0]
        
        #2)-------- first fetching the produciton data 

        production_data = ws[f"A{start}" : f"P{end}" ]
        rows_prod = [ [ value.value for value   in row ] for row in production_data ]

        d_prod = pd.DataFrame(rows_prod[1:] )

        d_prod[0] = d_prod[0].astype(str).str.strip()

        # add the date column
        report_date = ws["K3"].value
        d_prod.insert(0,"Date", report_date)

        # handle the PFS and the Avg.DesalterTemp
        d_prod[12] = d_prod[12].iloc[1]
        d_prod[15] = d_prod[15].iloc[1]

        
        # handle the inbetween lines
        d_prod.dropna(subset=[1],inplace=True)
        
        # handle wells with no production data

        # remove the rows with nan production data [ gross production , oil production , water cut ]
        rows_nan_prod_data = d_prod[[3,4,5]].isnull().all(axis=1)
        d_prod = d_prod[~rows_nan_prod_data]
        
        #3)second fetching the produciton data hours for each well
        hours_data = self.fetch_hours_data(ws,indcies)
        d_prod = pd.merge(d_prod, hours_data[[0,"hours","status","reason"]], on = 0, how ='inner') 

        d_prod.columns = [ "Date","Well" , "Jet Pump" , "Inj.Line Chock" , "Gross Production" , "Bs&w" , "Est. Oil Prod" , "Cum.Oil yesterday",
                    "Cum.Oil today" , "Injection Rate" , "WHIP" , "Flow Line (Psi)" , "Oil API" , "PFS (Bar)" , "Gas Prod (MSCF/Day)" ,
                    "GOR (SCF/BBL)" , "Avg. Desalter Temp C." , "hours" , "Status" , "Remarks" ]
        replace_dic = {
            "Muzhil-1": "M1_A_NK",
            "Muzhil-2": "M2_A_MT",
            "Muzhil-3D" : "M3_D_MT",
            "Muzhil-5" : "M5_A_NK",
            "Muzhil-9" : "M9_A_MT"
        }
        d_prod["Well"] = d_prod["Well"].replace(replace_dic)
        
        # process the hours data for any mistake
        d_prod["hours"] =  d_prod["hours"].apply( self.handle_hours_format)
        return d_prod


    def fetch_shipping_data( self, ws, indices):
        # define the table range
        start = indices["Cumulative"][0]
        end = start + 7

        # Get the table correspoding the production hours data 
        Cum_data = ws[f"A{start}":f"G{end}"]
        rows_cum = [[value.value for value   in row] for row in Cum_data]
        Cum_data = pd.DataFrame(rows_cum[1:], columns=rows_cum[0])
        Cum_data.dropna(how='all', axis=1, inplace=True)
        report_date = ws["K3"].value
        Cum_data.insert(0,"Date", report_date)
        Cum_data.columns = range(len(Cum_data.columns))
        Cum_data[1] = Cum_data[1].astype(str)
        # remove the rows with nan production data [ gross production , oil production , water cut ]
        rows_nan_prod_data = Cum_data[[2,3,4]].isnull().all(axis=1)
        Cum_data = Cum_data[~rows_nan_prod_data]

        return Cum_data

    def fetch_tank_data(self, ws, indices):
        # define the table range
        start = indices["Storage Tanks"][-1]
        # end = indices["Close Stock"][0]
        end = start + 6
        
        # Get the table correspoding the production hours data 
        tank_data = ws[f"A{start}":f"K{end}"]
        rows_tank = [[value.value for value   in row] for row in tank_data]
        d_tank = pd.DataFrame(rows_tank[1:], columns=rows_tank[0])
        report_date = ws["K3"].value
        d_tank.insert(0,"Date", report_date)
        d_tank.columns = range(len(d_tank.columns))
        d_tank.drop(columns=[2,3],inplace=True)
        rows_nan_prod_data = d_tank[[4,5,6]].isnull().all(axis=1)
        d_tank = d_tank[~rows_nan_prod_data]
        d_tank.columns= ["Date","Storage Tanks","Tank 09","Tank 01","Tank 04","Tank 07","Tank 02","Tank 05","Total" ,"Remarks"]

        # process the data for database insertion
        dic_map = {
            "Openning Volume  ":"Opening Stock",
            'Opening Volume  ':"Opening Stock",
            'Opening Stock' :'Opening Stock',
            'Total oil Received From MOPU':'Total Oil Received From MOPU'
        }
        d_tank["Storage Tanks"] = d_tank["Storage Tanks"].replace(dic_map)
        melted_df = pd.melt(d_tank.iloc[:,:-1] , id_vars = ["Date","Storage Tanks"], var_name="Tank Name" )
        result_df = melted_df.pivot(index=['Date', 'Tank Name'], columns='Storage Tanks', values='value').reset_index().reset_index(drop=True)
        result_df.index.name = "index"

        # get only the total values
        result_df= result_df[result_df["Tank Name"]=="Total"]

        return result_df

    def fetch_HPS_data(self, ws, indices):
        
        # define the table range
        start = indices["Freq."][0]
        end = start+2
        
        # Get the table correspoding the production hours data 
        HPS_data = ws[f"H{start}":f"N{end}"]
        rows_hps = [[value.value for value   in row] for row in HPS_data]
        HPS_data = pd.DataFrame(rows_hps[1:], columns=rows_hps[0])
        HPS_data.dropna(how='all', axis=1, inplace=True)
        HPS_data.rename(columns={None:"HPS Number"},inplace=True)
        report_date = ws["K3"].value
        HPS_data.insert(0,"Date", report_date)
        HPS_data.columns = range(len(HPS_data.columns))
        rows_nan_prod_data = HPS_data[[2,3,4]].isnull().all(axis=1)
        HPS_data = HPS_data[~rows_nan_prod_data]

        return HPS_data

    def fetch_test_data(self, ws, indices):
        # define the table range
        start = indices["Wells"][1]
        end = indices["Wells"][-1]
        
        # Get the table correspoding the production hours data 
        Test_data = ws[f"A{start}":f"F{end}"]
        rows_Test = [[value.value for value   in row] for row in Test_data]
        Test_data = pd.DataFrame(rows_Test[1:], columns=rows_Test[0])
        Test_data.dropna(how='all', axis=1, inplace=True)
        Test_data.rename(columns={None:"HPS Number"},inplace=True)
        # report_date = ws["K3"].value
        # Test_data.insert(0,"Date", report_date)
        Test_data.columns = range(len(Test_data.columns))
        Test_data = Test_data[Test_data[0].isin(self.Wells.Short_Name)]

        # process the data 
        Test_data.columns = ["Well","TestDate","GrossReturn","NetOil","WC","Gross.WC"]
        Test_data["WC"] = Test_data["WC"] * 100
        Test_data["Gross.WC"] = Test_data["Gross.WC"] * 100
        Test_data["TestDate"] = pd.to_datetime(Test_data["TestDate"])

        # return columns in order
        Test_data = Test_data[["TestDate","Well","GrossReturn","NetOil","WC","Gross.WC"]]
        replace_dic_test = {
            "M-1": "M1_A_NK",
            "M-2": "M2_A_MT",
            "M-3D" : "M3_D_MT",
            "M-5" : "M5_A_NK",
            "M-9" : "M9_A_MT"
        }
        Test_data["Well"] = Test_data["Well"].replace(replace_dic_test)

        return Test_data
    
    def fetch_all_data(self):

        wb =  openpyxl.load_workbook(self.file_path, read_only=True,data_only=True) 
        try:
            ws = wb['PZN Daily Production Report ']
        except:
            ws = wb['PZN Daily Production Report']

        # get the data
        indcies = self.get_tables_ranges(ws)
        prod_data = self.fetch_production_data(ws,indcies)
        test_data = self.fetch_test_data(ws,indcies)
        hps_data = self.fetch_HPS_data(ws,indcies)
        tank_data = self.fetch_tank_data(ws, indcies)
        shipping_data = self.fetch_shipping_data(ws, indcies)

        wb.close()

        print("Data is fetched successuly.")
        return prod_data,test_data,hps_data,tank_data,shipping_data
    


if __name__ == "__main__":

    file_path = r"C:\Users\asayed\Desktop\Email_Fetching\PZN_Daily_Production__Report 03-10-2023.xlsx".replace("\\","/")
    fetcher = OneFileFetch(file_path)
    prod_data,test_data,hps_data,tank_data,shipping_data = fetcher.fetch_all_data()
    print(prod_data)
    print(tank_data)
    print(test_data)
    print(hps_data)
    print(shipping_data)